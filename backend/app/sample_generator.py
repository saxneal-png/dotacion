import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SAMPLES_DIR = os.path.join(os.path.dirname(__file__), "..", "samples")
os.makedirs(SAMPLES_DIR, exist_ok=True)

def create_sample_files():
    # 1. Arturo Merino SLEP
    f1_path = os.path.join(SAMPLES_DIR, "Planilla Dotación Arturo Merino SLEP.xlsx")
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "Dotacion 2026"
    
    ws1["A1"] = "SERVICIO LOCAL DE EDUCACIÓN PÚBLICA VALLE DIGUILLÍN"
    ws1["A1"].font = Font(bold=True, size=12)
    ws1["A2"] = "ESTABLECIMIENTO: ESCUELA ARTURO MERINO BENÍTEZ"
    ws1["A2"].font = Font(bold=True)
    ws1["A3"] = "RBD: 17822-4"
    ws1["C3"] = "MATRÍCULA: 345"

    headers1 = ["RUT", "NOMBRE DOCENTE", "FUNCIÓN / CARGO / ASIGNATURA", "HORAS ASIGNADAS", "TOTAL HORAS CONTRATO"]
    for i, h in enumerate(headers1, 1):
        c = ws1.cell(row=5, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

    teachers_f1 = [
        ("12.345.678-9", "CARLOS ANDRÉS VALENZUELA MUÑOZ", "Director", 44, 44),
        ("14.567.890-1", "MARÍA TERESA LÓPEZ GARRIDO", "Inspectora General", 44, 44),
        ("15.678.901-2", "JUAN PABLO BUSTOS SILVA", "Jefe UTP", 44, 44),
        ("16.789.012-3", "ANDREA CAROLINA ROJAS PÉREZ", "Coordinadora PIE", 20, 44),
        ("16.789.012-3", "ANDREA CAROLINA ROJAS PÉREZ", "Atención PIE en aula", 24, 44),
        ("17.890.123-4", "FRANCISCO JAVIER MORALES RIVAS", "Lenguaje y Comunicación", 38, 44),
        ("17.890.123-4", "FRANCISCO JAVIER MORALES RIVAS", "Planificación", 6, 44),
        ("18.901.234-5", "CAMILA BEATRIZ SOTO CARVAJAL", "Matemática", 36, 40),
        ("18.901.234-5", "CAMILA BEATRIZ SOTO CARVAJAL", "Taller de Ajedrez", 4, 40),
        ("19.012.345-6", "RODRIGO ALEJANDRO FUENTES DÍAZ", "Ciencias Naturales", 32, 38),
        ("19.012.345-6", "RODRIGO ALEJANDRO FUENTES DÍAZ", "Comunidades CAP", 6, 38),
        ("13.456.789-0", "LORETO PAZ HENRÍQUEZ VEGA", "Encargada CRA", 22, 30),
        ("13.456.789-0", "LORETO PAZ HENRÍQUEZ VEGA", "Taller de Lectura", 8, 30),
        ("15.234.567-8", "GONZALO IGNACIO CASTILLO REYES", "Educación Física y Salud", 38, 42),
        ("15.234.567-8", "GONZALO IGNACIO CASTILLO REYES", "Taller Extraescolar Fútbol", 4, 42),
    ]

    for row_idx, t in enumerate(teachers_f1, 6):
        ws1.cell(row=row_idx, column=1, value=t[0])
        ws1.cell(row=row_idx, column=2, value=t[1])
        ws1.cell(row=row_idx, column=3, value=t[2])
        ws1.cell(row=row_idx, column=4, value=t[3])
        ws1.cell(row=row_idx, column=5, value=t[4])

    wb1.save(f1_path)

    # 2. República de México 2026
    f2_path = os.path.join(SAMPLES_DIR, "PLANILLA DE PRUEBA República de México 2026.xlsx")
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Dotacion_Mexico"

    ws2["A1"] = "PLANILLA DE DOTACIÓN Y CARGA HORARIA 2026"
    ws2["A2"] = "Establecimiento: ESCUELA REPÚBLICA DE MÉXICO"
    ws2["A3"] = "RBD: 16540"
    ws2["C3"] = "Matrícula: 280"

    headers2 = ["DOCENTE", "RUT", "ACTIVIDAD PEDAGOGICA / ADMINISTRATIVA", "HORAS LECTIVAS / NO LECTIVAS"]
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(row=5, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")

    teachers_f2 = [
        ("PATRICIO HERNÁN SALAZAR OPAZO", "11.222.333-4", "Directora", 44),
        ("CLAUDIA LORENA AGUILERA PARRA", "13.444.555-6", "Subdirectora", 40),
        ("ESTEBAN MAURICIO VEGA PINTO", "15.666.777-8", "Apoyo UTP", 30),
        ("ESTEBAN MAURICIO VEGA PINTO", "15.666.777-8", "Historia y Geografía", 14),
        ("MARCELA VIVIANA SEPÚLVEDA LEAL", "16.888.999-0", "Inglés", 36),
        ("MARCELA VIVIANA SEPÚLVEDA LEAL", "16.888.999-0", "Coordinación Enlaces", 8),
        ("DIEGO ALONSO RIQUELME CAMPOS", "17.111.222-3", "Artes Visuales", 28),
        ("DIEGO ALONSO RIQUELME CAMPOS", "17.111.222-3", "Taller SEP Pintura", 6),
        ("DIEGO ALONSO RIQUELME CAMPOS", "17.111.222-3", "Horas no lectivas", 4),
        ("FERNANDA ISABEL TORRES GÓMEZ", "18.333.444-5", "Trabajo Colaborativo PIE", 12),
        ("FERNANDA ISABEL TORRES GÓMEZ", "18.333.444-5", "Codocencia", 28),
    ]

    for row_idx, t in enumerate(teachers_f2, 6):
        ws2.cell(row=row_idx, column=1, value=t[0])
        ws2.cell(row=row_idx, column=2, value=t[1])
        ws2.cell(row=row_idx, column=3, value=t[2])
        ws2.cell(row=row_idx, column=4, value=t[3])

    wb2.save(f2_path)

    # 3. Pueblo Seco Santa Clara 2026
    f3_path = os.path.join(SAMPLES_DIR, "Planilla Dotación Pueblo Seco Santa Clara 2026.xlsx")
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    ws3.title = "Dotacion Docente"

    ws3["B2"] = "COLEGIO PUEBLO SECO SANTA CLARA"
    ws3["B3"] = "RBD: 18455"
    ws3["D3"] = "MATRÍCULA: 510"

    headers3 = ["N°", "NOMBRE PROFESOR", "RUT", "CARGO / ROL", "HORAS TOTALES"]
    for i, h in enumerate(headers3, 1):
        c = ws3.cell(row=5, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="B45309", end_color="B45309", fill_type="solid")

    teachers_f3 = [
        (1, "JAIME ALBERTO NAVARRETE VÁSQUEZ", "10.987.654-3", "Encargado de Escuela", 44),
        (2, "BEATRIZ ELENA CÁCERES BRAVO", "12.876.543-2", "Inspector General", 44),
        (3, "VÍCTOR MANUEL ZÚÑIGA SANDOVAL", "14.765.432-1", "Coord. Convivencia Escolar", 30),
        (4, "VÍCTOR MANUEL ZÚÑIGA SANDOVAL", "14.765.432-1", "Orientación frente a estudiantes", 14),
        (5, "DANIELA CONSTANZA VERGARA PINO", "16.654.321-0", "Educación Parvularia", 38),
        (6, "DANIELA CONSTANZA VERGARA PINO", "16.654.321-0", "Preparación de clases", 6),
        (7, "MATÍAS NICOLÁS OSSES LEIVA", "18.543.210-9", "Música", 30),
        (8, "MATÍAS NICOLÁS OSSES LEIVA", "18.543.210-9", "Taller Banda Escolar", 10),
        (9, "CAROLINA ANDREA MONTECINOS RUZ", "19.432.109-8", "Encargado SEP", 20),
        (10, "CAROLINA ANDREA MONTECINOS RUZ", "19.432.109-8", "Monitoreo de Cursos", 20),
    ]

    for row_idx, t in enumerate(teachers_f3, 6):
        ws3.cell(row=row_idx, column=1, value=t[0])
        ws3.cell(row=row_idx, column=2, value=t[1])
        ws3.cell(row=row_idx, column=3, value=t[2])
        ws3.cell(row=row_idx, column=4, value=t[3])
        ws3.cell(row=row_idx, column=5, value=t[4])

    wb3.save(f3_path)

    # 4. Liceo Yungay PROYECCIÓN 2027
    f4_path = os.path.join(SAMPLES_DIR, "Planilla Dotación Liceo Yungay PROYECCIÓN 2027.xlsx")
    wb4 = openpyxl.Workbook()
    ws4 = wb4.active
    ws4.title = "Yungay 2027"

    ws4["A1"] = "LICEO BICENTENARIO DE YUNGAY - PROYECCIÓN 2027"
    ws4["A2"] = "RBD: 19102"
    ws4["C2"] = "Alumnos Matrícula: 680"

    headers4 = ["DOCENTE", "ESPECIALIDAD / ASIGNATURA", "HORAS"]
    for i, h in enumerate(headers4, 1):
        c = ws4.cell(row=4, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")

    teachers_f4 = [
        ("ALEJANDRA MARISOL FUENTES VILLEGAS", "Rector", 44),
        ("FELIPE HERNANDO MORA SAAVEDRA", "Jefe UTP", 44),
        ("GABRIEL ENRIQUE ACUÑA BARRERA", "Filosofía", 36),
        ("GABRIEL ENRIQUE ACUÑA BARRERA", "Coordinación Departamento Lenguaje", 8),
        ("ISIDORA PAZ MIRANDA CONTRERAS", "Biología", 34),
        ("ISIDORA PAZ MIRANDA CONTRERAS", "Curriculista", 10),
        ("SEBASTIÁN ANDRÉS TOLEDO LAGOS", "Física y Química", 38),
        ("SEBASTIÁN ANDRÉS TOLEDO LAGOS", "Coordinación Medio Ambiente", 6),
        ("VALERIA ANDREA CORTÉS URREA", "Aula de Recursos", 35),
        ("VALERIA ANDREA CORTÉS URREA", "Trabajo Colaborativo", 9),
    ]

    for row_idx, t in enumerate(teachers_f4, 5):
        ws4.cell(row=row_idx, column=1, value=t[0])
        ws4.cell(row=row_idx, column=2, value=t[1])
        ws4.cell(row=row_idx, column=3, value=t[2])

    wb4.save(f4_path)

    return [f1_path, f2_path, f3_path, f4_path]

if __name__ == "__main__":
    paths = create_sample_files()
    print(f"Created {len(paths)} sample files.")
