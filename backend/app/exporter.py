import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any

def generate_consolidated_excel(schools: List[Dict[str, Any]], detailed_teachers: List[Dict[str, Any]] = None) -> bytes:
    wb = openpyxl.Workbook()
    
    # ==========================
    # Sheet 1: CONSOLIDADO SLEP
    # ==========================
    ws = wb.active
    ws.title = "Consolidado SLEP"
    ws.views.sheetView[0].showGridLines = True

    # Title Banner
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "SERVICIO LOCAL DE EDUCACIÓN PÚBLICA (SLEP) - CONSOLIDADO INSTITUCIONAL DE DOTACIÓN DOCENTE"
    title_cell.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Subtitle with date/source
    ws.merge_cells("A2:G2")
    sub_cell = ws["A2"]
    sub_cell.value = "Clasificación oficial de Horas Docentes: Aula, Directivas y Técnicas"
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color="334155")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Table Headers
    headers = [
        "RBD",
        "ESTABLECIMIENTO",
        "MATRÍCULA",
        "HORAS DOCENTES AULA",
        "HORAS DOCENTES DIRECTIVAS",
        "HORAS DOCENTES TÉCNICAS",
        "TOTAL HORAS EE"
    ]

    header_row = 4
    ws.row_dimensions[header_row].height = 28

    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )
    thick_bottom = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="medium", color="1E3A8A")
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thick_bottom

    # Data Rows
    current_row = 5
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for s in schools:
        ws.row_dimensions[current_row].height = 22
        fill = alt_fill if current_row % 2 == 0 else white_fill

        c1 = ws.cell(row=current_row, column=1, value=s.get("rbd", ""))
        c1.alignment = Alignment(horizontal="center", vertical="center")
        
        c2 = ws.cell(row=current_row, column=2, value=s.get("establishment", ""))
        c2.alignment = Alignment(horizontal="left", vertical="center")

        c3 = ws.cell(row=current_row, column=3, value=s.get("matricula", 0))
        c3.alignment = Alignment(horizontal="right", vertical="center")
        c3.number_format = "#,##0"

        c4 = ws.cell(row=current_row, column=4, value=round(s.get("horas_aula", 0.0), 1))
        c4.alignment = Alignment(horizontal="right", vertical="center")
        c4.number_format = "#,##0.0"

        c5 = ws.cell(row=current_row, column=5, value=round(s.get("horas_directivas", 0.0), 1))
        c5.alignment = Alignment(horizontal="right", vertical="center")
        c5.number_format = "#,##0.0"

        c6 = ws.cell(row=current_row, column=6, value=round(s.get("horas_tecnicas", 0.0), 1))
        c6.alignment = Alignment(horizontal="right", vertical="center")
        c6.number_format = "#,##0.0"

        # Formula for total: D + E + F
        c7 = ws.cell(row=current_row, column=7, value=f"=SUM(D{current_row}:F{current_row})")
        c7.alignment = Alignment(horizontal="right", vertical="center")
        c7.number_format = "#,##0.0"
        c7.font = Font(name="Calibri", bold=True)

        for col in range(1, 8):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = fill
            cell.border = thin_border

        current_row += 1

    # Totals Row
    ws.row_dimensions[current_row].height = 25
    totals_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    totals_font = Font(name="Calibri", size=11, bold=True, color="0F172A")
    totals_border = Border(
        top=Side(style="thin", color="0F172A"),
        bottom=Side(style="double", color="0F172A"),
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1")
    )

    ws.cell(row=current_row, column=1, value="")
    t_label = ws.cell(row=current_row, column=2, value="TOTAL GENERAL SLEP")
    t_label.alignment = Alignment(horizontal="left", vertical="center")

    t_mat = ws.cell(row=current_row, column=3, value=f"=SUM(C5:C{current_row-1})")
    t_mat.number_format = "#,##0"
    t_mat.alignment = Alignment(horizontal="right", vertical="center")

    t_aula = ws.cell(row=current_row, column=4, value=f"=SUM(D5:D{current_row-1})")
    t_aula.number_format = "#,##0.0"
    t_aula.alignment = Alignment(horizontal="right", vertical="center")

    t_dir = ws.cell(row=current_row, column=5, value=f"=SUM(E5:E{current_row-1})")
    t_dir.number_format = "#,##0.0"
    t_dir.alignment = Alignment(horizontal="right", vertical="center")

    t_tec = ws.cell(row=current_row, column=6, value=f"=SUM(F5:F{current_row-1})")
    t_tec.number_format = "#,##0.0"
    t_tec.alignment = Alignment(horizontal="right", vertical="center")

    t_tot = ws.cell(row=current_row, column=7, value=f"=SUM(G5:G{current_row-1})")
    t_tot.number_format = "#,##0.0"
    t_tot.alignment = Alignment(horizontal="right", vertical="center")

    for col in range(1, 8):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = totals_fill
        cell.font = totals_font
        cell.border = totals_border

    # Adjust column widths
    col_widths = {1: 14, 2: 44, 3: 16, 4: 24, 5: 26, 6: 25, 7: 20}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ==========================
    # Sheet 2: DETALLE DOCENTE
    # ==========================
    if detailed_teachers:
        ws_det = wb.create_sheet(title="Detalle Docente y Auditoría")
        ws_det.views.sheetView[0].showGridLines = True

        det_headers = [
            "RBD",
            "ESTABLECIMIENTO",
            "RUT",
            "DOCENTE",
            "ACTIVIDAD / ASIGNATURA",
            "HORAS",
            "CLASIFICACIÓN",
            "ORIGEN REGLA / IA"
        ]

        ws_det.row_dimensions[1].height = 26
        for col_idx, header in enumerate(det_headers, 1):
            cell = ws_det.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        d_row = 2
        for t in detailed_teachers:
            ws_det.row_dimensions[d_row].height = 20
            ws_det.cell(row=d_row, column=1, value=t.get("rbd", ""))
            ws_det.cell(row=d_row, column=2, value=t.get("establishment", ""))
            ws_det.cell(row=d_row, column=3, value=t.get("rut", ""))
            ws_det.cell(row=d_row, column=4, value=t.get("teacher_name", ""))
            ws_det.cell(row=d_row, column=5, value=t.get("activity", ""))
            
            c_h = ws_det.cell(row=d_row, column=6, value=t.get("hours", 0.0))
            c_h.number_format = "#,##0.0"

            c_cat = ws_det.cell(row=d_row, column=7, value=t.get("category", ""))
            if t.get("category") == "AULA":
                c_cat.font = Font(color="1E40AF", bold=True)
            elif t.get("category") == "DIRECTIVA":
                c_cat.font = Font(color="B91C1C", bold=True)
            elif t.get("category") == "TECNICA":
                c_cat.font = Font(color="047857", bold=True)

            ws_det.cell(row=d_row, column=8, value=t.get("source", ""))
            d_row += 1

        det_widths = {1: 14, 2: 38, 3: 15, 4: 35, 5: 35, 6: 12, 7: 18, 8: 25}
        for col_idx, width in det_widths.items():
            ws_det.column_dimensions[get_column_letter(col_idx)].width = width

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
