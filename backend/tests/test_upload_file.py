import io
import openpyxl
import requests

def test_upload():
    # Create an in-memory workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dotacion"
    ws["A1"] = "COLEGIO DE PRUEBA SLEP"
    ws["A2"] = "RBD: 19999"
    ws["C2"] = "Matrícula: 320"

    ws.append([])
    ws.append(["RUT", "DOCENTE", "CARGO / ASIGNATURA", "HORAS"])
    ws.append(["11.111.111-1", "Docente Prueba 1", "Director", 44])
    ws.append(["22.222.222-2", "Docente Prueba 2", "Lenguaje", 38])
    ws.append(["22.222.222-2", "Docente Prueba 2", "Planificación", 6])
    ws.append(["33.333.333-3", "Docente Prueba 3", "Coord. PIE", 20])
    ws.append(["33.333.333-3", "Docente Prueba 3", "Atención PIE en aula", 24])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    files = [
        ("files", ("Planilla_Prueba.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))
    ]

    print("Sending POST /api/upload ...")
    res = requests.post("http://127.0.0.1:8080/api/upload", files=files)
    print(f"Status: {res.status_code}")
    if res.status_code != 200:
        print("Response error:", res.text)
    assert res.status_code == 200, f"Expected 200 OK, got {res.status_code}"
    
    data = res.json()
    print("Schools returned:", len(data["schools"]))
    print("Teachers returned:", len(data["teachers"]))
    print("KPIs:", data["kpis"])
    assert len(data["schools"]) == 1
    assert data["schools"][0]["rbd"] == "19999"
    print("\n>>> TEST UPLOAD PASSED WITH 200 OK! <<<")

if __name__ == "__main__":
    test_upload()
