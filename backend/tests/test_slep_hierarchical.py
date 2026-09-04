import io
import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import openpyxl
from app.excel_parser import parse_excel_file
from app.classifier import classify_activity

def test_hierarchical():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dotacion SLEP"

    ws["A1"] = "ESCUELA EJEMPLO SLEP"
    ws["A2"] = "RBD: 17822"
    ws["C2"] = "Matrícula: 345"

    headers = [
        "NOMBRES", "RUN", "HORAS CONTRATO", "TITULARES SUB GRAL", "CONTRATA SUB GRAL",
        "TITULARES SEP", "CONTRATA SEP", "TITULARES PIE", "CONTRATA PIE",
        "PK", "K", "1", "2", "3", "4", "5", "6", "7", "8", "1°", "2°", "3°", "4°",
        "TOTAL HA*", "TOTAL HC** Sub. Gral", "TOTAL HC** Sub. SEP", "TOTAL HC** Sub. PIE", "TOTAL HC"
    ]
    ws.append(headers)

    # Teacher master row
    row_doc = ["SEPULVEDA CERDA, CAROLINA OLGA", "10.274.857-3", 44, 37, 0, 7, 0, 0, 0,
               None, None, None, None, None, None, 4, 4, 6, 6, None, None, None, None,
               None, 37, 7, 0, 44]
    ws.append(row_doc)

    # Activity rows
    def make_row(name, ha=None, hc_gral=None, hc_sep=None, hc_pie=None, tot_hc=None):
        r = [name] + [None]*22 + [ha, hc_gral, hc_sep, hc_pie, tot_hc]
        return r

    ws.append(make_row("Historia", ha=16, hc_gral="12:00"))
    ws.append(make_row("Formación Ciudadana", ha=4, hc_gral="3:00"))
    ws.append(make_row("Tiempo Funciones no lectivas (artº 69)", hc_gral="3:00"))
    ws.append(make_row("Encargada CRA", hc_gral="14:00", hc_sep="7:00"))
    ws.append(make_row("Recreos 60/40"))
    ws.append(make_row("Recreo 65/35", hc_gral="1:34"))
    ws.append(make_row("Horas no Lectivas 60/40"))
    ws.append(make_row("Horas No lectivas 65/35", hc_gral="6:26"))
    ws.append(make_row("Total", hc_gral="40:00", hc_sep="7:00", hc_pie="0:00", tot_hc="47:00"))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    parsed = parse_excel_file(buf.getvalue(), "Planilla_Carolina.xlsx")
    print(f"RBD: {parsed['rbd']}, Establecimiento: {parsed['establishment']}")
    print(f"Total actividades extraídas: {len(parsed['teachers'])}")

    assert len(parsed["teachers"]) > 0, "Debe extraer actividades del docente"

    h_aula = 0.0
    h_tec = 0.0
    h_dir = 0.0

    for t in parsed["teachers"]:
        cat, src = classify_activity(t["activity"])
        print(f"  - [{cat}] {t['activity']}: {t['hours']} hrs ({src})")
        if cat == "AULA":
            h_aula += t["hours"]
        elif cat == "TECNICA":
            h_tec += t["hours"]
        elif cat == "DIRECTIVA":
            h_dir += t["hours"]

    print(f"\nResumen: Aula={h_aula:.2f}h, Técnica={h_tec:.2f}h, Directiva={h_dir:.2f}h, Total={h_aula+h_tec+h_dir:.2f}h")

    assert h_aula == 15.0, f"Horas aula esperadas: 15.0, obtenidas: {h_aula}"
    assert h_tec == 32.0, f"Horas técnicas esperadas: 32.0, obtenidas: {h_tec}"
    assert h_dir == 0.0, f"Horas directivas esperadas: 0.0, obtenidas: {h_dir}"
    print("\n>>> ¡PRUEBA DE ESTRUCTURA JERÁRQUICA SLEP PASÓ CON 100% DE EXACTITUD! <<<")

if __name__ == "__main__":
    test_hierarchical()
