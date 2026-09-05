import sys
import os
import io
import datetime
import pandas as pd
import openpyxl

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from app.excel_parser import parse_hours_cell, parse_excel_file, is_rut, looks_like_teacher_name
import app.sample_generator as sg

def test_parse_hours_numbers():
    assert parse_hours_cell(44) == 44.0
    assert parse_hours_cell(44.0) == 44.0
    assert parse_hours_cell('44') == 44.0
    assert parse_hours_cell('44.5') == 44.5
    assert parse_hours_cell('44,5') == 44.5

def test_parse_hours_time_strings():
    assert parse_hours_cell('44:00') == 44.0
    assert parse_hours_cell('[44]:00') == 44.0
    assert parse_hours_cell('12:30') == 12.5
    assert parse_hours_cell('01:34:00') == 1.57
    assert parse_hours_cell('1:34') == 1.57

def test_parse_hours_native_python_objects():
    # datetime.time
    assert parse_hours_cell(datetime.time(12, 0)) == 12.0
    assert parse_hours_cell(datetime.time(1, 34)) == 1.57

    # datetime.timedelta
    assert parse_hours_cell(datetime.timedelta(days=1, seconds=72000)) == 44.0
    assert parse_hours_cell(datetime.timedelta(seconds=45000)) == 12.5

    # datetime.datetime (Excel epoch offsets for time)
    assert parse_hours_cell(datetime.datetime(1899, 12, 30, 12, 30)) == 12.5
    assert parse_hours_cell(datetime.datetime(1900, 1, 1, 14, 0)) == 14.0
    assert parse_hours_cell(datetime.datetime(1900, 1, 2, 20, 0)) == 44.0
    assert parse_hours_cell(pd.Timestamp('1900-01-01 12:30:00')) == 12.5

def test_parse_hours_text_with_metadata_and_years():
    # Strings that previously caused 442026.0 or 441520.0 hours
    assert parse_hours_cell('44 HRS (SEP 2026)') == 44.0
    assert parse_hours_cell('44 HORAS') == 44.0
    assert parse_hours_cell('30 hrs.') == 30.0
    assert parse_hours_cell('20 horas semanales') == 20.0
    assert parse_hours_cell('ORD 1234 44 HRS') == 44.0
    assert parse_hours_cell('30 HORAS LEY 19.070') == 30.0

def test_parse_hours_dates_and_sanity_limits():
    # Dates should return 0.0 (not hours)
    assert parse_hours_cell(datetime.date(2024, 3, 1)) == 0.0
    assert parse_hours_cell(datetime.datetime(2024, 3, 1, 0, 0)) == 0.0
    assert parse_hours_cell('01/03/2024') == 0.0
    assert parse_hours_cell('2024-03-01') == 0.0

    # Values exceeding sanity bounds (> 50h) return 0.0
    assert parse_hours_cell(2024) == 0.0
    assert parse_hours_cell(19070) == 0.0
    assert parse_hours_cell('19070') == 0.0
    assert parse_hours_cell(1520) == 0.0

def test_is_rut():
    assert is_rut('12.345.678-9')
    assert is_rut('12345678-9')
    assert is_rut('12345678-k')
    assert is_rut('12345678-K')
    assert is_rut('12345678K')
    assert is_rut('15432198')  # pure numeric
    assert not is_rut('')
    assert not is_rut('Lenguaje')
    assert not is_rut('44')

def test_sample_arturo_merino_no_omitted_teachers():
    sg.create_sample_files()
    f = os.path.join(sg.SAMPLES_DIR, 'Planilla Dotación Arturo Merino SLEP.xlsx')
    parsed = parse_excel_file(f, 'Planilla Dotación Arturo Merino SLEP.xlsx')

    # In previous code, Arturo Merino extracted 0 teachers!
    assert len(parsed['teachers']) == 15
    for t in parsed['teachers']:
        assert 0 < t['hours'] <= 44.0
        assert t['teacher_name'] != ''
        assert t['rut'] != ''
        assert t['activity'] != ''

def test_sample_republica_mexico():
    f = os.path.join(sg.SAMPLES_DIR, 'PLANILLA DE PRUEBA República de México 2026.xlsx')
    parsed = parse_excel_file(f, 'PLANILLA DE PRUEBA República de México 2026.xlsx')
    assert len(parsed['teachers']) == 11
    for t in parsed['teachers']:
        assert 0 < t['hours'] <= 44.0

def test_sample_liceo_yungay():
    f = os.path.join(sg.SAMPLES_DIR, 'Planilla Dotación Liceo Yungay PROYECCIÓN 2027.xlsx')
    parsed = parse_excel_file(f, 'Planilla Dotación Liceo Yungay PROYECCIÓN 2027.xlsx')
    assert len(parsed['teachers']) == 10
    for t in parsed['teachers']:
        assert 0 < t['hours'] <= 44.0

def test_sample_pueblo_seco():
    f = os.path.join(sg.SAMPLES_DIR, 'Planilla Dotación Pueblo Seco Santa Clara 2026.xlsx')
    parsed = parse_excel_file(f, 'Planilla Dotación Pueblo Seco Santa Clara 2026.xlsx')
    assert len(parsed['teachers']) == 10
    for t in parsed['teachers']:
        assert 0 < t['hours'] <= 44.0

def test_merged_cells_teacher_block():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['RBD: 99999', 'ESCUELA CON CELDAS COMBINADAS'])
    ws.append(['RUT', 'DOCENTE', 'ACTIVIDAD', 'HORAS ASIGNADAS', 'TOTAL HORAS CONTRATO'])
    ws.append(['17.890.123-4', 'FRANCISCO MORALES', 'Lenguaje', 38, 44])
    ws.append([None, None, 'Planificación', 6, 44])
    ws.append(['18.901.234-5', 'CAMILA SOTO', 'Matemática', 40, 40])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    parsed = parse_excel_file(buf.getvalue(), 'Merged_Cells_Test.xlsx')
    assert len(parsed['teachers']) == 3
    t1 = parsed['teachers'][0]
    t2 = parsed['teachers'][1]
    t3 = parsed['teachers'][2]

    assert t1['teacher_name'] == 'FRANCISCO MORALES'
    assert t1['activity'] == 'Lenguaje'
    assert t1['hours'] == 38.0

    assert t2['teacher_name'] == 'FRANCISCO MORALES'
    assert t2['activity'] == 'Planificación'
    assert t2['hours'] == 6.0

    assert t3['teacher_name'] == 'CAMILA SOTO'
    assert t3['activity'] == 'Matemática'
    assert t3['hours'] == 40.0

def test_day_fraction_hours():
    # Excel time serial fractions (SheetJS raw number fallback)
    assert parse_hours_cell(0.267361) == 6.42
    assert parse_hours_cell(0.5) == 12.0
    assert parse_hours_cell('0.267361') == 6.42
    assert parse_hours_cell('0.5') == 12.0
    assert parse_hours_cell(0.125) == 3.0

def test_teacher_without_rut_not_omitted():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['RBD: 11111', 'ESCUELA PRUEBA SIN RUT'])
    ws.append(['NOMBRES', 'RUN', 'ACTIVIDAD', 'HORAS TOTALES'])
    ws.append(['DOCENTE POR CONTRATAR', '', 'Matemática', 30])
    ws.append(['VACANTE LENGUAJE', 'S/RUT', 'Lenguaje', 44])
    ws.append(['MARIA PEREZ (DIRECTOR)', '12.345.678-9', '', 44])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    parsed = parse_excel_file(buf.getvalue(), 'Test_Sin_Rut.xlsx')
    assert len(parsed['teachers']) == 3
    assert parsed['teachers'][0]['teacher_name'] == 'DOCENTE POR CONTRATAR'
    assert parsed['teachers'][0]['activity'] == 'Matemática'
    assert parsed['teachers'][0]['hours'] == 30.0

    assert parsed['teachers'][1]['teacher_name'] == 'VACANTE LENGUAJE'
    assert parsed['teachers'][1]['activity'] == 'Lenguaje'
    assert parsed['teachers'][1]['hours'] == 44.0

    assert parsed['teachers'][2]['teacher_name'] == 'MARIA PEREZ'
    assert parsed['teachers'][2]['activity'] == 'Director'
    assert parsed['teachers'][2]['hours'] == 44.0

def test_subtotal_rows_not_counted_as_hours():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['RBD: 22222', 'ESCUELA SUB-TOTALES'])
    ws.append(['NOMBRES', 'RUN', 'ACTIVIDAD', 'TOTAL HC'])
    ws.append(['PEDRO PASCAL', '15.111.222-3', 'Física', 20])
    ws.append(['PEDRO PASCAL', '15.111.222-3', 'Química', 24])
    ws.append(['TOTAL DOCENTE', '', '', 44])
    ws.append(['TOTAL GENERAL', '', '', 44])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    parsed = parse_excel_file(buf.getvalue(), 'Test_Subtotales.xlsx')
    assert len(parsed['teachers']) == 2
    total_h = sum(t['hours'] for t in parsed['teachers'])
    assert total_h == 44.0  # Exactly 20 + 24 = 44, not 44 + 44 + 44 = 132!
